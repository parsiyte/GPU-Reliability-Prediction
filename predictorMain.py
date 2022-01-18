import pandas as pd
import numpy as np
import csv
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.linear_model import SGDClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import make_pipeline
import xlsxwriter  
import openpyxl  
from pathlib import Path  # path

from sklearn.ensemble import RandomForestRegressor
from sklearn import svm
from sklearn.ensemble import GradientBoostingRegressor
import getopt, sys # for taking argument from the command line 


def classifier(model, metrics, target):
    train_input = []
    test_input = []
    train_target = []
    test_target = []
    prediction = []

    for i in range(0, len(target)):
        for j in range (0, len(target)):
            if i == j:
                test_input.append(metrics[j])
                test_target.append(target[j])
            else:
                train_input.append(metrics[j])
                train_target.append(target[j])

        model.fit(train_input, train_target)
        prediction.append(model.predict(test_input))
        train_input = []
        test_input = []
        train_target = []
        test_target = []
    return prediction

def class_identifier_1(error, fault_type):
    rates = []
    if fault_type == 'crash':    
        for i in range(0, len(error)):
            if error[i] < (0.105/2):
                rates.append(0)
            else:
                rates.append(1)
    else:
        for i in range(0, len(error)):
            if error[i] < 0.05:
                rates.append(0)
            else:
                rates.append(1)        
    return rates

def class_identifier_2(error, fault_type):
    rates = []
    if fault_type == 'crash':
        for i in range(0, len(error)):
            if error[i] < (0.105/3):
                rates.append(0)
            elif error[i] >= (0.105/3) and error[i] < (0.105*2/3):
                    rates.append(1)
            else:
                    rates.append(2)
    else:
        for i in range(0, len(error)):
            if error[i] < 0.03:
                rates.append(0)
            elif error[i] >= 0.03 and error[i] < 0.07:
                    rates.append(1)
            else:
                    rates.append(2)         
    return rates

def error_calculator_classification(test, pred_rf, pred_sgd, pred_gb, apps, fault_rates, fault_type, num_class):

    acc_rf = 0
    acc_gb = 0 
    acc_sgd = 0 

    for i in range (len(test)):
        if pred_rf[i] == test[i]:
            acc_rf += 1
        if pred_sgd[i] == test[i]:
            acc_sgd += 1
        if pred_gb[i] == test[i]:
            acc_gb += 1

    print("random forrest acc: " + "{:.3f}".format((float(acc_rf)/float(len(test)))*100) )
    print("ensemble total acc: " + "{:.3f}".format((float(acc_sgd)/float(len(test)))*100))
    print("gradient boosting acc: " + "{:.3f}".format((float(acc_gb)/float(len(test)))*100))

    if (num_class == 2) and fault_type == 'sdc':
        workbook = xlsxwriter.Workbook("SDC_ERR_Classifications_with_" + str(num_class) + "_classes.xlsx")
        sheet = workbook.add_worksheet()

    if (num_class == 3) and fault_type == 'sdc':
        workbook = xlsxwriter.Workbook("SDC_ERR_Classifications_with_" + str(num_class) + "_classes.xlsx")
        sheet = workbook.add_worksheet()

    if (num_class == 2) and fault_type == 'crash':
        workbook = xlsxwriter.Workbook("Crash_ERR_Classifications_with_" + str(num_class) + "_classes.xlsx")
        sheet = workbook.add_worksheet()

    if (num_class == 3) and fault_type == 'crash':
        workbook = xlsxwriter.Workbook("Crash_ERR_Classifications_with_" + str(num_class) + "_classes.xlsx")
        sheet = workbook.add_worksheet()


    sheet.write(0, 0, 'Applications')
    sheet.write(0, 1, 'SDC Error Rates')
    sheet.write(0, 2, 'Actual Class')
    sheet.write(0, 3, 'Predicted Class(RF)')
    sheet.write(0, 4, 'Predicted Class(ensemble)')
    sheet.write(0, 5, 'Predicted Class(GB)')

    for i in range(0,len(pred_sgd)):
        sheet.write(i+1, 0, apps[i])
        sheet.write(i+1, 1, fault_rates[i])
        sheet.write(i+1, 2, test[i])
        for j in range (3):
            if j == 0:
                sheet.write(i+1, j+3, pred_rf[i])
            if j == 1:
                sheet.write(i+1, j+3, pred_sgd[i])
            if j == 2:
                sheet.write(i+1, j+3, pred_gb[i])
    sheet.write(len(pred_sgd)+1, 3, "{:.3f}".format((float(acc_rf)/float(len(test)))*100))
    sheet.write(len(pred_sgd)+1, 4, "{:.3f}".format((float(acc_sgd)/float(len(test)))*100))
    sheet.write(len(pred_sgd)+1, 5, "{:.3f}".format((float(acc_gb)/float(len(test)))*100))

    workbook.close()

def precision_recall(pred_class, actual_class, rate, num_class, classifier):

    if (num_class == 2):
        pred_sel_0 = 0
        pred_sel_1 = 0

        act_sel_0 = 0
        act_sel_1 = 0
        for i in range (0, len(pred_class)):
            if pred_class[i] == 0:
                pred_sel_0 += 1
            else:
                pred_sel_1 += 1

            if actual_class[i] == 0:
                act_sel_0 += 1
            else:
                act_sel_1 += 1

        true_sel_0 = 0
        true_sel_1 = 0

        for i in range (0, len(pred_class)):
            if pred_class[i] == 0 and pred_class[i] == actual_class[i]:
                true_sel_0 += 1
            elif pred_class[i] == 1 and pred_class[i] == actual_class[i]:
                true_sel_1 += 1

        precision_0 = float(true_sel_0)/float(pred_sel_0)
        recall_0 = float(true_sel_0)/float(act_sel_0)

        precision_1 = float(true_sel_1)/float(pred_sel_1)
        recall_1 = float(true_sel_1)/float(act_sel_1)

        workbook = xlsxwriter.Workbook("SDC_Precision_Recall_F-Score_2Classes_"+ str(classifier) +".xlsx")
        sheet = workbook.add_worksheet()

        sheet.write(0, 0, 'Class')
        sheet.write(0, 1, 'Precison')
        sheet.write(0, 2, 'Recall')
        sheet.write(0, 3, 'F-score')

        if precision_0 + recall_0 != 0:
            sheet.write(1, 0, 0)
            sheet.write(1, 1, "{:.3f}".format(precision_0))
            sheet.write(1, 2, "{:.3f}".format(recall_0))
            sheet.write(1, 3, "{:.3f}".format(2*(precision_0*recall_0)/(precision_0 + recall_0)))

        if precision_1 + recall_1 != 0:
            sheet.write(2, 0, 1)
            sheet.write(2, 1, "{:.3f}".format(precision_1))
            sheet.write(2, 2, "{:.3f}".format(recall_1))
            sheet.write(2, 3, "{:.3f}".format(2*(precision_1*recall_1)/(precision_1 + recall_1)))

        workbook.close()


    elif (num_class==3):
        pred_sel_0 = 0
        pred_sel_1 = 0
        pred_sel_2 = 0

        act_sel_0 = 0
        act_sel_1 = 0
        act_sel_2 = 0

        for i in range (0, len(pred_class)):
            if pred_class[i] == 0:
                pred_sel_0 += 1
            elif pred_class[i] == 1:
                pred_sel_1 += 1
            else:
                pred_sel_2 += 1 

            if actual_class[i] == 0:
                act_sel_0 += 1
            elif actual_class[i] == 1:
                act_sel_1 += 1
            else:
                act_sel_2 += 1

        true_sel_0 = 0
        true_sel_1 = 0
        true_sel_2 = 0

        for i in range (0, len(pred_class)):
            if pred_class[i] == 0 and pred_class[i] == actual_class[i]:
                true_sel_0 += 1
            elif pred_class[i] == 1 and pred_class[i] == actual_class[i]:
                true_sel_1 += 1
            elif pred_class[i] == 2 and pred_class[i] == actual_class[i]:
                true_sel_2 += 1

        workbook = xlsxwriter.Workbook("SDC_Precision_Recall_F-Score_3Classes_"+ str(classifier) +".xlsx")
        sheet = workbook.add_worksheet()

        sheet.write(0, 0, 'Class')
        sheet.write(0, 1, 'Precison')
        sheet.write(0, 2, 'Recall')
        sheet.write(0, 3, 'F-score')

        if pred_sel_0 != 0 and act_sel_0 != 0:
            precision_0 = float(true_sel_0)/float(pred_sel_0)
            recall_0 = float(true_sel_0)/float(act_sel_0)

            if recall_0 + precision_0!= 0:
                sheet.write(1, 0, 0)
                sheet.write(1, 1, "{:.3f}".format(precision_0))
                sheet.write(1, 2, "{:.3f}".format(recall_0))
                sheet.write(1, 3, "{:.3f}".format(2*(precision_0*recall_0)/(precision_0 + recall_0)))

        if pred_sel_1 != 0 and act_sel_1 != 0:
            precision_1 = float(true_sel_1)/float(pred_sel_1)
            recall_1 = float(true_sel_1)/float(act_sel_1)

            if recall_1 + precision_1 != 0:
                sheet.write(2, 0, 1)
                sheet.write(2, 1, "{:.3f}".format(precision_1))
                sheet.write(2, 2, "{:.3f}".format(recall_1))
                sheet.write(2, 3, "{:.3f}".format(2*(precision_1*recall_1)/(precision_1 + recall_1)))

        if pred_sel_2 != 0 and act_sel_2 != 0:
            precision_2 = float(true_sel_2)/float(pred_sel_2)
            recall_2 = float(true_sel_2)/float(act_sel_2)

            if recall_2 + precision_2 != 0:
                sheet.write(3, 0, 2)
                sheet.write(3, 1, "{:.3f}".format(precision_2))
                sheet.write(3, 2, "{:.3f}".format(recall_2))
                sheet.write(3, 3, "{:.3f}".format(2*(precision_2*recall_2)/(precision_2 + recall_2)))

        workbook.close()


def classifier_initialization(features, fault_classes, apps, fault_rates, fault_type, num_class):
    model = RandomForestClassifier(max_depth=2, random_state=0)
    pred_rf = classifier(model, features, fault_classes)
    
    model = make_pipeline(StandardScaler(), SGDClassifier(max_iter=1000, tol=1e-3))
    pred_sgd = classifier(model, features, fault_classes)
    
    model = GradientBoostingClassifier(n_estimators=20, max_depth=2, random_state=0)
    pred_gb = classifier(model, features, fault_classes)

    error_calculator_classification(fault_classes, pred_rf, pred_sgd, pred_gb, apps, fault_rates, fault_type, num_class)
    
    precision_recall(pred_rf, fault_classes, fault_rates, num_class, "rf")
    precision_recall(pred_sgd, fault_classes, fault_rates, num_class, "sgd")
    precision_recall(pred_gb, fault_classes, fault_rates, num_class, "gb")
   
def regressor_model(features, label, model_name):
    prediction = np.zeros((4, len(features)), dtype=float)
    y_train = []
    x_train = []
    x_test = []

    for j in range(0, 4):
        if model_name == 'GB':
            model = GradientBoostingRegressor(learning_rate=0.001)
            if j == 1:
                model = GradientBoostingRegressor(learning_rate=0.01)
            if j == 2:
                model = GradientBoostingRegressor(learning_rate=0.1)
            if j == 3:
                model = GradientBoostingRegressor(learning_rate=0.25)
        
        elif model_name == 'RF':
            model = RandomForestRegressor(n_estimators=100, random_state=20)
            if j == 1:
                model = RandomForestRegressor(n_estimators=1000, random_state=20)
            if j == 2:
                model = RandomForestRegressor(n_estimators=100, random_state=50)
            if j == 3:
                model = RandomForestRegressor(n_estimators=10000, random_state=100)
     
        else:
            model = svm.SVR(kernel='poly', degree=2)
            if j == 1:
                model = svm.SVR(kernel='poly', degree=3)
            if j == 2:
                model = svm.SVR(kernel='rbf')
            if j == 3:
                model = svm.SVR(kernel='sigmoid')

        for i in range(0, len(features)):
            x_test = np.array([features[i], features[i]])

            a = features[0:i, :]
            b = features[i + 1:len(features), :]
            x_train = np.concatenate((a, b))

            a = label[0:i]
            b = label[i + 1:len(features)]
            y_train = np.concatenate((a, b))

            model.fit(x_train, y_train)
            x = model.predict(x_test)
            prediction[j][i] = x[0]

            y_train = []
            x_train = []
            x_test = []
    return prediction


def error_prints_regressor_preds(gb_pred, rf_pred, svm_pred, fault_rates, apps, fault_type, feature_type):

    acc_rf = np.zeros((4), dtype=float)
    acc_gb = np.zeros((4), dtype=float) 
    acc_svm = np.zeros((4), dtype=float)

    for j in range (4):
        for i in range (len(fault_rates)):
            acc_rf[j] += (1-(abs(rf_pred[j][i]-fault_rates[i])/fault_rates[i]))
            acc_gb[j] += (1-(abs(gb_pred[j][i]-fault_rates[i])/fault_rates[i]))
            acc_svm[j] += (1-(abs(svm_pred[j][i]-fault_rates[i])/fault_rates[i]))

        print("random forrest average acc: " + "{:.3f}".format((float(acc_rf[j])/float(len(fault_rates)))*100) )
        print("svm average acc: " + "{:.3f}".format((float(acc_svm[j])/float(len(fault_rates)))*100))
        print("gradient boosting average acc: " + "{:.3f}".format((float(acc_gb[j])/float(len(fault_rates)))*100))

    if fault_type == 'masked':
        if feature_type == 'all_features':
            workbook = xlsxwriter.Workbook("MaskedFaultsPredictionWithRegression.xlsx")
        else:
            workbook = xlsxwriter.Workbook("MaskedFaultsPredictionWithRegression_Selected_features.xlsx")
    if fault_type == 'crash':
        if feature_type == 'all_features':
            workbook = xlsxwriter.Workbook("CrashFaultsPredictionWithRegression.xlsx")
        else:
            workbook = xlsxwriter.Workbook("CrashFaultsPredictionWithRegression_Selected_features.xlsx")
    if fault_type == 'SDC':
        if feature_type == 'all_features':
            workbook = xlsxwriter.Workbook("SDCFaultsPredictionWithRegression.xlsx")
        else:
            workbook = xlsxwriter.Workbook("SDCFaultsPredictionWithRegression_Selected_features.xlsx")

    sheet = workbook.add_worksheet()
    sheet.write(0, 0, 'Applications')
    sheet.write(0, 1, 'fault_rate ' + fault_type)
    sheet.write(0, 2, 'prediction_1_svm')
    sheet.write(0, 3, 'prediction_2_svm')
    sheet.write(0, 4, 'prediction_3_svm')
    sheet.write(0, 5, 'prediction_4_svm')
    sheet.write(0, 6, 'prediction_1_rf')
    sheet.write(0, 7, 'prediction_2_rf')
    sheet.write(0, 8, 'prediction_3_rf')
    sheet.write(0, 9, 'prediction_4_rf')
    sheet.write(0, 10, 'prediction_1_gb')
    sheet.write(0, 11, 'prediction_2_gb')
    sheet.write(0, 12, 'prediction_3_gb')
    sheet.write(0, 13, 'prediction_4_gb')

    for i in range(0,len(fault_rates)):
        sheet.write(i+1, 0, apps[i])
        sheet.write(i+1, 1, fault_rates[i])
        for j in range (4):
            sheet.write(i+1, j+2, svm_pred[j][i])
            sheet.write(i+1, j+6, rf_pred[j][i])
            sheet.write(i+1, j+10, gb_pred[j][i])

    for j in range (4):
        sheet.write(len(fault_rates)+1, 2+j, "{:.3f}".format((float(acc_svm[j])/float(len(fault_rates)))*100))
        sheet.write(len(fault_rates)+1, 6+j, "{:.3f}".format((float(acc_rf[j])/float(len(fault_rates))))*100)
        sheet.write(len(fault_rates)+1, 10+j, "{:.3f}".format((float(acc_gb[j])/float(len(fault_rates)))*100))

    workbook.close()


def regressor_initialization(features, fault_rates, apps, fault_type, feature_type):
    gb_pred = regressor_model(features, fault_rates, 'GB')
    rf_pred = regressor_model(features, fault_rates, 'RF')
    svm_pred = regressor_model(features, fault_rates, 'SVM')

    error_prints_regressor_preds(gb_pred, rf_pred, svm_pred, fault_rates, apps, fault_type, feature_type)


metrics_GPGPUsim = pd.read_excel('data_metrics_GPUsim.xls')
data_GPGPUsim = pd.DataFrame(metrics_GPGPUsim,
                    columns=['Applications', 'Load_instruction', 'Store_instruction', 'Param_Mem_instruction', 'Total_instruction', 'IPC', 'Sim_Rate',
                    'Global_Mem_Read', 'Global_Mem_Write', 'BW_Utilization', 'Warp_Occupancy', 'Control_Flow_Inst_Intensity', 'Data_Mov_Inst_Intensity',
                    'Float_Point_Inst_Intensity',    'Integer_Arithmetic_Inst_Intensity', 'Logical_Inst_Intensity',    'Load_Inst_Intensity', 
                    'Predicate_Inst_Intensity', 'SDC', 'Crash', 'Masked'])

features_GPGPUsim = data_GPGPUsim[['Load_instruction', 'Store_instruction', 'Param_Mem_instruction', 'Total_instruction', 'IPC', 'Sim_Rate',
                    'Global_Mem_Read', 'Global_Mem_Write', 'BW_Utilization', 'Warp_Occupancy', 'Control_Flow_Inst_Intensity', 'Data_Mov_Inst_Intensity',
                    'Float_Point_Inst_Intensity',   'Integer_Arithmetic_Inst_Intensity', 'Logical_Inst_Intensity',  'Load_Inst_Intensity', 
                    'Predicate_Inst_Intensity']].values

features_sel_sdc_GPGPUsim = data_GPGPUsim[['Param_Mem_instruction', 'IPC', 'Sim_Rate', 'Data_Mov_Inst_Intensity', 'Float_Point_Inst_Intensity', 
                    'Logical_Inst_Intensity',  'Load_Inst_Intensity', 'Predicate_Inst_Intensity']].values

features_sel_crash_GPGPUsim = data_GPGPUsim[['Load_instruction', 'Store_instruction', 'Param_Mem_instruction', 'Total_instruction', 'Global_Mem_Read', 
                    'Global_Mem_Write', 'BW_Utilization', 'Warp_Occupancy', 'Control_Flow_Inst_Intensity', 'Data_Mov_Inst_Intensity',
                    'Integer_Arithmetic_Inst_Intensity', 'Load_Inst_Intensity']].values

features_sel_masked_GPGPUsim = data_GPGPUsim[['Load_instruction', 'Store_instruction', 'Total_instruction', 'Global_Mem_Read', 'Global_Mem_Write', 'Warp_Occupancy', 
                    'Float_Point_Inst_Intensity',   'Integer_Arithmetic_Inst_Intensity', 'Logical_Inst_Intensity',  'Predicate_Inst_Intensity']].values

metrics_Nsight = pd.read_excel('data_metrics_NSC.xls')
data_Nsight = pd.DataFrame(metrics_Nsight,
					columns=['Applications', "SOL_SM", "SOL_MEM", "SOL_L1_Tex_Cache", "SOL_L2_Cache", "SOL_DRAM", "Duration", 
                    "Elapsed_Cycle", "IPC", "Mem_Throughput", "L2_Hit_Rate", "Mem_Busy", "Max_Band", "Active_Warp_Per_Sch", 
                    "Warp_Cyc_Inst", "Executed_Inst", "Reg_Per_Thread", "Achieved_Occupancy", "Achieved_Active_Warp", "SDC", "Crash", "Masked"])

features_Nsight = data_Nsight[["SOL_SM", "SOL_MEM", "SOL_L1_Tex_Cache", "SOL_L2_Cache", "SOL_DRAM", "Duration", 
                    "Elapsed_Cycle", "IPC", "Mem_Throughput", "L2_Hit_Rate", "Mem_Busy", "Max_Band", "Active_Warp_Per_Sch", 
                    "Warp_Cyc_Inst", "Executed_Inst", "Reg_Per_Thread", "Achieved_Occupancy", "Achieved_Active_Warp"]].values

apps_GPGPUsim = data_GPGPUsim['Applications'].values
masked_GPGPUsim = data_GPGPUsim['Masked'].values
sdc_GPGPUsim = data_GPGPUsim['SDC'].values
crash_GPGPUsim = data_GPGPUsim['Crash'].values

apps_Nsight = data_Nsight["Applications"].values
masked_Nsight = data_Nsight["Masked"].values
sdc_Nsight = data_Nsight["SDC"].values
crash_Nsight = data_Nsight["Crash"].values

features_sel_sdc_Nsight = data_Nsight[["SOL_DRAM", "IPC", "Active_Warp_Per_Sch", 
                                       "Warp_Cyc_Inst", "Reg_Per_Thread", "Achieved_Occupancy", "Achieved_Active_Warp"]].values

features_sel_masked_Nsight = data_Nsight[["SOL_SM", "SOL_L1_Tex_Cache", "SOL_L2_Cache", "L2_Hit_Rate", "Max_Band", "Active_Warp_Per_Sch", 
                                          "Reg_Per_Thread", "Achieved_Occupancy", "Achieved_Active_Warp"]].values

features_sel_crash_Nsight = data_Nsight[["SOL_SM", "SOL_MEM", "SOL_L1_Tex_Cache", "SOL_L2_Cache", "SOL_DRAM", "Mem_Throughput", "L2_Hit_Rate", 
                                         "Mem_Busy", "Warp_Cyc_Inst",  "Reg_Per_Thread"]].values


argumentList = sys.argv[1:]

##############################
#### GPGPUSim Experiments ####
##############################

#### SDC with all features ####
if argumentList[0] == '--gpgpu-sim' and argumentList[1] == '--sdc' and argumentList[2] == '--all_features' :
    print('results for 2 classes')
    classifier_initialization(features_GPGPUsim, class_identifier_1(sdc_GPGPUsim, 'sdc'), apps_GPGPUsim, sdc_GPGPUsim, 'sdc', 2)
    print('results for 3 classes')
    classifier_initialization(features_GPGPUsim, class_identifier_2(sdc_GPGPUsim, 'sdc'), apps_GPGPUsim, sdc_GPGPUsim, 'sdc', 3)
#### SDC with selected features ####
elif argumentList[0] == '--gpgpu-sim' and argumentList[1] == '--sdc' and argumentList[2] == '--sel_features' :
    print('results for 2 classes')
    classifier_initialization(features_sel_sdc_GPGPUsim, class_identifier_1(sdc_GPGPUsim, 'sdc'), apps_GPGPUsim, sdc_GPGPUsim, 'sdc', 2)
    print('results for 3 classes')
    classifier_initialization(features_sel_sdc_GPGPUsim, class_identifier_2(sdc_GPGPUsim, 'sdc'), apps_GPGPUsim, sdc_GPGPUsim, 'sdc', 3)

#### CRASH with all features ####
elif argumentList[0] == '--gpgpu-sim' and argumentList[1] == '--crash' and argumentList[2] == '--all_features' :
    print('results for 2 classes')
    classifier_initialization(features_GPGPUsim, class_identifier_1(crash_GPGPUsim, 'crash'), apps_GPGPUsim, crash_GPGPUsim, 'crash', 2)
    print('results for 3 classes')
    classifier_initialization(features_GPGPUsim, class_identifier_2(crash_GPGPUsim, 'crash'), apps_GPGPUsim, crash_GPGPUsim, 'crash', 3)
#### CRASH with selected features ####
elif argumentList[0] == '--gpgpu-sim' and argumentList[1] == '--crash' and argumentList[2] == '--sel_features' :
    print('results for 2 classes')
    classifier_initialization(features_sel_crash_GPGPUsim, class_identifier_1(crash_GPGPUsim, 'crash'), apps_GPGPUsim, crash_GPGPUsim, 'crash', 2)
    print('results for 3 classes')
    classifier_initialization(features_sel_crash_GPGPUsim, class_identifier_2(crash_GPGPUsim, 'crash'), apps_GPGPUsim, crash_GPGPUsim, 'crash', 3)

#### Regression Experiments ####
elif argumentList[0] == '--gpgpu-sim' and argumentList[1] == '--masked':
    #### Masked Regression ####
    print("Masked fault prediction results obtained from the all features")
    regressor_initialization(features_GPGPUsim, masked_GPGPUsim, apps_GPGPUsim, 'masked', 'all_features')
    print("Masked fault prediction results obtained from the selected features")
    regressor_initialization(features_sel_masked_GPGPUsim, masked_GPGPUsim, apps_GPGPUsim, 'masked', 'sel_features')

##############################
##### Nsight Experiments #####
##############################

#### SDC with all features ####
elif argumentList[0] == '--nsight-compute' and argumentList[1] == '--sdc' and argumentList[2] == '--all_features' :
    print('results for 2 classes')
    classifier_initialization(features_Nsight, class_identifier_1(sdc_Nsight, 'sdc'), apps_Nsight, sdc_Nsight, 'sdc', 2)
    print('results for 3 classes')
    classifier_initialization(features_Nsight, class_identifier_2(sdc_Nsight, 'sdc'), apps_Nsight, sdc_Nsight, 'sdc', 3)
#### SDC with selected features ####
elif argumentList[0] == '--nsight-compute' and argumentList[1] == '--sdc' and argumentList[2] == '--sel_features' :
    print('results for 2 classes')
    classifier_initialization(features_sel_sdc_Nsight, class_identifier_1(sdc_Nsight, 'sdc'), apps_Nsight, sdc_Nsight, 'sdc', 2)
    print('results for 3 classes')
    classifier_initialization(features_sel_sdc_Nsight, class_identifier_2(sdc_Nsight, 'sdc'), apps_Nsight, sdc_Nsight, 'sdc', 3)

#### CRASH with all features ####
elif argumentList[0] == '--nsight-compute' and argumentList[1] == '--crash' and argumentList[2] == '--all_features' :
    print('results for 2 classes')
    classifier_initialization(features_Nsight, class_identifier_1(sdc_Nsight, 'crash'), apps_Nsight, crash_Nsight, 'crash', 2)
    print('results for 3 classes')
    classifier_initialization(features_Nsight, class_identifier_2(sdc_Nsight, 'crash'), apps_Nsight, crash_Nsight, 'crash', 3)
#### CRASH with selected features ####
elif argumentList[0] == '--nsight-compute' and argumentList[1] == '--crash' and argumentList[2] == '--sel_features' :
    print('results for 2 classes')
    classifier_initialization(features_sel_crash_Nsight, class_identifier_1(crash_Nsight, 'crash'), apps_Nsight, crash_Nsight, 'crash', 2)
    print('results for 3 classes')
    classifier_initialization(features_sel_crash_Nsight, class_identifier_2(crash_Nsight, 'crash'), apps_Nsight, crash_Nsight, 'crash', 3)

#### Regression Experiments ####
elif argumentList[0] == '--nsight-compute' and argumentList[1] == '--masked':
    #### Masked Regression ####
    print("Masked fault prediction results obtained from the all features")
    regressor_initialization(features_Nsight, masked_Nsight, apps_Nsight, 'masked', 'all_features')
    print("Masked fault prediction results obtained from the selected features")
    regressor_initialization(features_sel_masked_Nsight, masked_Nsight, apps_Nsight, 'masked', 'sel_features')

elif argumentList[0] == '--nsight-compute' and argumentList[1] == '--others':
    #### SDC Regression ####
    print("SDC prediction results obtained from the all features")
    regressor_initialization(features_Nsight, sdc_Nsight, apps_Nsight, 'SDC', 'all_features')
    print("SDC prediction results obtained from the selected features")
    regressor_initialization(features_sel_sdc_Nsight, sdc_Nsight, apps_Nsight, 'SDC', 'sel_features')

    #### CRASH regression ####
    print("Crash Fault prediction results obtained from the all features")
    regressor_initialization(features_Nsight, crash_Nsight, apps_Nsight, 'crash', 'all_features')
    print("Crash Fault prediction results obtained from the selected features")
    regressor_initialization(features_sel_crash_Nsight, crash_Nsight, apps_Nsights, 'crash', 'sel_features')

elif argumentList[0] == '--gpgpu-sim' and argumentList[1] == '--others':
    #### SDC Regression ####
    print("SDC prediction results obtained from the all features")
    regressor_initialization(features_GPGPUsim, sdc_GPGPUsim, apps_GPGPUsim, 'SDC', 'all_features')
    print("SDC prediction results obtained from the selected features")
    regressor_initialization(features_sel_sdc_GPGPUsim, sdc_GPGPUsim, apps_GPGPUsim, 'SDC', 'sel_features')

    #### CRASH regression ####
    print("Crash Fault prediction results obtained from the all features")
    regressor_initialization(features_GPGPUsim, crash_GPGPUsim, apps_GPGPUsim, 'crash', 'all_features')
    print("Crash Fault prediction results obtained from the selected features")
    regressor_initialization(features_sel_crash_GPGPUsim, crash_GPGPUsim, apps_GPGPUsim, 'crash', 'sel_features')
